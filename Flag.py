import openpyxl
from openpyxl import *
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
source_file = load_workbook("insert file")
sheet = source_file["insert file"]
ft = Font(color="008000")


#CHANGES

# -----------JULY 8th 6:30 ADDED FUNCTION TO VIEW BOTH ON EDITNAMES()

"""TOD DO'S


-Fix loop bug that identifies every chosen cell

-add color change

-ADD A LAST UPDATED

--ADD VALIDATION TO EVERY CHOICE LOOP

-SET IT TO RUN FOR ENTIRE DOCUMENT


"""

#FLAG ON NAMES
##############################################################################################################

#Flags Based on name of employees 
def flagBasedOnNames():

    #Menu options for user
    moreNames = True
    
    while moreNames != False and moreNames != False:
        nameToBeFlagged = input("Type in the name of the employee you would like to flag (NOTE: CASE SENSITIVE):")
        
        #Begins looping through excel doc
        for i in range(1, 20):
            word = sheet.cell(row=i, column=4).value

         

            #EVALUATE IF IT CONTAINS THE NAME
            result = word.find(nameToBeFlagged)
            if result != -1:

                

                #if it contains the desired name than select the far right most cell to put in the Flag "Y"
                IinString = str(i)
                finalstring = 'Q' + IinString

                
                #checks if there are any other flags so as not to delete them
                alreadyThere = sheet[finalstring].value
                

                #adds the new flag

                #checks to make sure that there is no value
                if alreadyThere == None:
                    
                    #if there isnt one it just assigns the flag
                     sheet[finalstring] = 'NAME'
                     print("All cells created by " + nameToBeFlagged + " have been flagged")
                else:

                    #if there is one it assigns the previous flags and the new flags
                    sheet[finalstring] = alreadyThere + ', NAME'
                    print("All cells created by " + nameToBeFlagged + " have been flagged")
                #Saves file
                source_file.save('JUNE302018.xlsx')


        #Menu Options
        correctInput = False
        while correctInput == False:
            choice = input("would you like to flag more names? if yes please type 'yes' or 'no'") 
            if choice == 'yes':
                moreNames = True
                correctInput = True
            elif choice == 'no':
                moreNames = False
                correctInput = True
            else:
                print ("please type yes or no as your answer")
                correctInput = False
        
            
           
        
    source_file.close()

#flagBasedOnNames()



    
#Flag on dates
###################################################################################################################

def flagOnDates():
    
        chosenDate = int(input("What year would you like to flag by? Please enter a 4 digit number only: "))
        #Loop through and find all dates run
        for i in range(2, 20):

            #Gets the year in the last run column of the cell and turns it into a string
            date = str(sheet.cell(row=i, column=8).value)

            #sepcifies it to the year and makes it an integer
            yearInt = int(date[0:4])

            #if the year is less than chosen year it flags it
            if yearInt < chosenDate:

                #assigns flag column
                IinString = str(i)
                finalstring = 'Q' + IinString

                    
                #checks if there are any other flags so as not to delete them
                alreadyThere = sheet[finalstring].value
                    

                #adds the new flag

                #checks to make sure that there is no value
                if alreadyThere == None:
                        
                    #if there isnt one it just assigns the flag
                    sheet[finalstring] = 'DATE'
                         
                else:

                    #if there is one it assigns the previous flags and the new flags
                    sheet[finalstring] = alreadyThere + ', DATE'

                #Saves file
                source_file.save('JUNE302018.xlsx')
        print("All dates less than " + str(chosenDate) + " have been flagged")
        
        source_file.close()

    #If date is less than 2019 then flag it



#Implement Tags on file
###################################################################################################################





#EDIT BASED ON FLAG TYPE


#MENU

def EditName():


    #Prompts user for their flag selection to ZZZ something
    print("Would you like to add ZZZ based on owner flags or created date flags or both?")
    flagChoice = input("Enter 1 for owner, 2 for Last run date or 3 for both:")

    #EMpty variably to assign the flag choce to
    flagFilter = ""

    #Decides Flag cgoice of name or date and puts it in variable FlagFIlter
    if flagChoice == 1:
        flagFilter = "NAME"
        for i in range(1, 20):
            word = sheet.cell(row=i, column=17).value

            #Looks for the chosen flag
        
            if word != None:

                result = word.find(flagFilter)

                if result != -1:
                    
                    

                    #if selected flag is not found then it selects column A
                    IinString = str(i)
                    finalstring = 'A' + IinString

                    colorString = sheet[finalstring]
                    colorString.font = ft

                    #Gathers what is already currently in that cell
                    alreadyThere = sheet[finalstring].value
                    
                    

                    #Adds ZZZ to it
                    sheet[finalstring] = "ZZZ" + alreadyThere

                    #Saves File
                    source_file.save('JUNE302018.xlsx')
                    source_file.close()
                    
    elif flagChoice == 3:
        
        for i in range(1, 20):
            word = sheet.cell(row=i, column=17).value

            #Looks for the chosen flag
        
            if word != None:

                result = word.find('NAME')
                result2 = word.find('DATE')
                

                if result != -1 and result2 != -1:

                    #if selected flag is not found then it selects column A
                    IinString = str(i)
                    finalstring = 'A' + IinString

                    #Gathers what is already currently in that cell
                    alreadyThere = sheet[finalstring].value

                    #Adds ZZZ to it
                    sheet[finalstring] = "ZZZ" + alreadyThere

                    #Saves File
                    source_file.save('JUNE302018.xlsx')
                    source_file.close()
        

        
    else:
        
        flagFilter = "DATE"
        
        for i in range(1, 20):
            word = sheet.cell(row=i, column=17).value

            #Looks for the chosen flag
        
            if word != None:

                result = word.find(flagFilter)

                if result != -1:

                    #if selected flag is not found then it selects column A
                    IinString = str(i)
                    finalstring = 'A' + IinString

                    #Gathers what is already currently in that cell
                    alreadyThere = sheet[finalstring].value

                    #Adds ZZZ to it
                    sheet[finalstring] = "ZZZ" + alreadyThere

                    #Saves File
                    source_file.save('JUNE302018.xlsx')
                    source_file.close()
        
    
    print("ZZZ was added to selected flag files")
            

            
#######################################################################################################################################################################







#MAIN MENU
#######################################################

print("Welcome to the Report Auto Flagger. This program can automatically parse")
print("through excel sheets of reports and flag the reports \nbased on last run date and names of the owner.\n")
print("\nIt can then alter the name of the report to \nZZZ so that it will come up last when being sorted")
print("\n")

done = False
while done != True:
    choiceOfProcess = int(input("Would you like to flag based on name (press 1) \nor flag based on date (press 2) \nor alter name (press 3)?:\n"))

    while choiceOfProcess > 3 or choiceOfProcess <= 0:
        choiceOfProcess = int(input("Invalid answer, please enter a valid choice: \n would you like to flag based on name (press 1) or flag based on date (press 2) or alter name (press 3)?:\n"))
    if choiceOfProcess == 1:
        flagBasedOnNames()
    elif choiceOfProcess == 2:
        flagOnDates()
    elif choiceOfProcess == 3:
        EditName()
    doneChoice = int(input("would you like to start another process? Please enter 1 for yes or 2 for no: \n"))
    while doneChoice > 2 or doneChoice <= 0:
        doneChoice = int(input("error, invalid choice, please enter either a 1 for yes or a 2 for no:\n"))
    if doneChoice == 1:
        done = False
    if doneChoice == 2:
        done = True

    
        

     
        




















